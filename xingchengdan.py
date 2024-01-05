#!/usr/bin/env python
# -*- coding:utf-8 -*-
# author: qi xy 
# date: 2023/06/02    time: 22:14 
# filename：baoxiao.py
# -*- coding: utf-8 -*-

import time

print("当前系统时间为："+time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) )
timerecord = []
current_time = str(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
timerecord.append(current_time)
print(timerecord)

import zipfile
import os
from turtle import shape

import PyPDF2
import pdfplumber
import shutil
from openpyxl import Workbook,load_workbook
from pathlib import Path

from zipfile import ZipFile


def support_gbk(zip_file: ZipFile):
    name_to_info = zip_file.NameToInfo
    # copy map first
    for name, info in name_to_info.copy().items():
        real_name = name.encode('cp437').decode('gbk')
        if real_name != name:
            info.filename = real_name
            del name_to_info[name]
            name_to_info[real_name] = info
    return zip_file


def get_file_from_zip(filename, filename_fz):
    #fanstasy_zip = zipfile.ZipFile(filename, 'r')  # 解压zip文件
    with support_gbk(ZipFile(filename,'r')) as zfp:
        # zfp.extractall(r'./中文不乱码')
      a1 = zfp.extract(filename_fz)  # 从zip文件中获得名为filename_fz的文件
      zfp.close()  # 关闭zip文件
      return a1


in_directory = os.path.dirname('E:\\SIAT\\baoxiao\\xingchengdan\\')

# Create output directory
out_directory = os.path.dirname('E:\\SIAT\\baoxiao\\output\\')
if not os.path.exists(out_directory):
    os.makedirs(out_directory)


files = os.listdir(in_directory)

output_excelName = 'new_pdf-jh1226.xlsx'
file_name_out = os.path.join(out_directory, output_excelName)
if not os.path.exists(file_name_out):
    wb = Workbook()     # 新建工作簿，默认带有一张表单为sheet
    wb.save(file_name_out)

for file in files:
        file_path = os.path.join(in_directory, file)

        if os.path.isfile(file_path):
            print(file)
        #选种后 按TAB 键缩进
        if file.endswith(".zip"):
            with zipfile.ZipFile(file_path, 'r') as zf:
                for fn in zf.namelist():
                    right_fn = fn.encode('cp437').decode('gbk')  # 将文件名正确编码
                    with open(right_fn, 'wb') as output_file:  # 创建并打开新文件
                        with zf.open(fn, 'r') as origin_file:  # 打开原文件
                            shutil.copyfileobj(origin_file, output_file)  # 将原文件内容复制到新文件
            file_name = get_file_from_zip(file_path,'滴滴出行行程报销单.pdf')
        elif file.endswith(".pdf"):
            file_name = file_path

        with pdfplumber.open(file_name) as p:
            for i in range(len(p.pages)):
                page0 = p.pages[0]
                table0 = page0.extract_table()
                #leaders_1.extend(leaders_2)
                # b=table[1:]
                # print(table[0:])
                # print(table[1:])
                if i>0:
                    page = p.pages[i]
                    table1 = page.extract_table()
                    #table2 = table0.append(table1[1:])
                    table2 = table0 + table1[1:]
            #data =[item.replace('\n','') for item in table0]
            for ii in range(len(table2)):
                for jj in range(len(table2[ii])):
                    table2[ii][jj] = table2[ii][jj].replace('\n','')
        data = table2
        workbook = load_workbook(filename=file_name_out)
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(filename = file_name_out)
